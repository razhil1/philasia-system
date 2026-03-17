from flask import Blueprint, render_template, flash, redirect, url_for, request, current_app, abort, jsonify, send_file, Response
from flask_login import login_required, current_user
from functools import wraps
from app import db
from app.models import (Item, Warehouse, ProjectSite, Stock, Movement, Request,
                         RequestItem, Category, User, AssetUnit, AssetUnitMovement,
                         MOVEMENT_TYPES, ITEM_TYPES, ASSET_STATUSES, ASSET_CONDITIONS)
from app.forms import (ItemForm, WarehouseForm, ProjectSiteForm, MovementForm,
                        RequestForm, RequestItemForm, UserForm, CategoryForm,
                        ApproveRequestForm, RejectRequestForm, ReportFilterForm,
                        ChangePasswordForm, AssetUnitForm, AssetMovementForm)
from sqlalchemy import func, desc, or_
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename
import os
import io

main = Blueprint('main', __name__)


def permission_required(perm):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.can(perm):
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('main.dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator


def _update_stock(item_id, location_type, warehouse_id, site_id, delta):
    if location_type in (None, 'none', 'external'):
        return
    q = Stock.query.filter_by(item_id=item_id, location_type=location_type)
    if location_type == 'warehouse':
        q = q.filter_by(warehouse_id=warehouse_id)
    else:
        q = q.filter_by(site_id=site_id)
    stock = q.first()
    if stock is None:
        stock = Stock(item_id=item_id, location_type=location_type,
                      warehouse_id=warehouse_id if location_type == 'warehouse' else None,
                      site_id=site_id if location_type == 'site' else None,
                      quantity=0)
        db.session.add(stock)
    stock.quantity = float(stock.quantity) + delta
    stock.last_updated = datetime.utcnow()


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@main.route('/')
@main.route('/dashboard')
@login_required
def dashboard():
    total_items = Item.query.filter_by(is_active=True).count()
    total_warehouses = Warehouse.query.filter_by(is_active=True).count()
    total_sites = ProjectSite.query.count()
    active_projects = ProjectSite.query.filter_by(status='active').count()
    pending_requests = Request.query.filter_by(status='pending').count()
    total_movements_today = Movement.query.filter(
        func.date(Movement.date) == date.today()).count()

    recent_movements = Movement.query.order_by(desc(Movement.date)).limit(8).all()
    recent_requests = Request.query.order_by(desc(Request.created_at)).limit(5).all()

    low_stock_items = []
    for item in Item.query.filter_by(is_active=True).all():
        total_qty = item.total_stock()
        if total_qty < float(item.reorder_level or 0):
            low_stock_items.append({'item': item, 'current_qty': total_qty})

    # Chart data: movements per day last 7 days
    chart_labels = []
    chart_inflow = []
    chart_outflow = []
    for i in range(6, -1, -1):
        d = date.today() - timedelta(days=i)
        chart_labels.append(d.strftime('%b %d'))
        inflow_types = ['delivery', 'return', 'pullout']
        outflow_types = ['transfer', 'site_transfer', 'consumption', 'scrap', 'adjustment']
        inflow = db.session.query(func.sum(Movement.quantity)).filter(
            func.date(Movement.date) == d,
            Movement.movement_type.in_(inflow_types)).scalar() or 0
        outflow = db.session.query(func.sum(Movement.quantity)).filter(
            func.date(Movement.date) == d,
            Movement.movement_type.in_(outflow_types)).scalar() or 0
        chart_inflow.append(float(inflow))
        chart_outflow.append(float(outflow))

    return render_template('inventory/dashboard.html',
        total_items=total_items, total_warehouses=total_warehouses,
        total_sites=total_sites, active_projects=active_projects,
        pending_requests=pending_requests, total_movements_today=total_movements_today,
        recent_movements=recent_movements, recent_requests=recent_requests,
        low_stock_items=low_stock_items[:5],
        chart_labels=chart_labels, chart_inflow=chart_inflow, chart_outflow=chart_outflow,
        now=datetime.now())


# ─── CATEGORIES ───────────────────────────────────────────────────────────────

@main.route('/categories')
@login_required
def category_list():
    categories = Category.query.order_by(Category.name).all()
    return render_template('inventory/category_list.html', categories=categories)


@main.route('/categories/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_categories')
def category_create():
    form = CategoryForm()
    if form.validate_on_submit():
        cat = Category(name=form.name.data, description=form.description.data)
        db.session.add(cat)
        db.session.commit()
        flash('Category created successfully.', 'success')
        return redirect(url_for('main.category_list'))
    return render_template('inventory/category_form.html', form=form, title='New Category')


@main.route('/categories/<int:cat_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('manage_categories')
def category_edit(cat_id):
    cat = Category.query.get_or_404(cat_id)
    form = CategoryForm(obj=cat)
    if form.validate_on_submit():
        cat.name = form.name.data
        cat.description = form.description.data
        db.session.commit()
        flash('Category updated.', 'success')
        return redirect(url_for('main.category_list'))
    return render_template('inventory/category_form.html', form=form, title='Edit Category')


# ─── ITEMS ────────────────────────────────────────────────────────────────────

@main.route('/items')
@login_required
def item_list():
    items = Item.query.filter_by(is_active=True).order_by(Item.name).all()
    item_stats = []
    for item in items:
        total_qty = item.total_stock()
        item_stats.append({'item': item, 'total_qty': total_qty, 'low': item.is_low_stock()})
    return render_template('inventory/item_list.html', item_stats=item_stats)


@main.route('/items/<int:item_id>')
@login_required
def item_detail(item_id):
    item = Item.query.get_or_404(item_id)
    stocks = Stock.query.filter_by(item_id=item_id).all()
    movements = Movement.query.filter_by(item_id=item_id).order_by(desc(Movement.date)).limit(20).all()
    return render_template('inventory/item_detail.html', item=item, stocks=stocks, movements=movements)


@main.route('/items/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_inventory')
def item_create():
    form = ItemForm()
    if form.validate_on_submit():
        filename = None
        if form.photo.data and form.photo.data.filename:
            filename = secure_filename(form.photo.data.filename)
            form.photo.data.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
        item = Item(sku=form.sku.data, name=form.name.data, description=form.description.data,
                    category_id=form.category_id.data, unit=form.unit.data,
                    unit_cost=form.unit_cost.data or 0,
                    reorder_level=form.reorder_level.data or 0,
                    item_type=form.item_type.data,
                    photo=filename, is_active=form.is_active.data)
        db.session.add(item)
        db.session.commit()
        flash('Item registered successfully.', 'success')
        return redirect(url_for('main.item_list'))
    return render_template('inventory/item_form.html', form=form, title='New Item')


@main.route('/items/<int:item_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('manage_inventory')
def item_edit(item_id):
    item = Item.query.get_or_404(item_id)
    form = ItemForm(obj=item)
    if form.validate_on_submit():
        if form.photo.data and form.photo.data.filename:
            filename = secure_filename(form.photo.data.filename)
            form.photo.data.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
            item.photo = filename
        item.sku = form.sku.data
        item.name = form.name.data
        item.description = form.description.data
        item.category_id = form.category_id.data
        item.unit = form.unit.data
        item.unit_cost = form.unit_cost.data or 0
        item.reorder_level = form.reorder_level.data or 0
        item.item_type = form.item_type.data
        item.is_active = form.is_active.data
        db.session.commit()
        flash('Item updated.', 'success')
        return redirect(url_for('main.item_detail', item_id=item_id))
    return render_template('inventory/item_form.html', form=form, title='Edit Item', item=item)


# ─── ASSET UNIT MANAGEMENT ────────────────────────────────────────────────────

@main.route('/items/<int:item_id>/units')
@login_required
def asset_unit_list(item_id):
    item = Item.query.get_or_404(item_id)
    if not item.is_asset:
        flash('This item is a consumable — it does not have individual unit tracking.', 'info')
        return redirect(url_for('main.item_detail', item_id=item_id))
    units = AssetUnit.query.filter_by(item_id=item_id).order_by(AssetUnit.status, AssetUnit.asset_tag).all()
    warehouse_list = Warehouse.query.filter_by(is_active=True).order_by('name').all()
    site_list = ProjectSite.query.order_by('name').all()
    warehouses = {w.id: w.name for w in warehouse_list}
    sites = {s.id: s.name for s in site_list}
    counts = item.asset_count_by_status()
    return render_template('inventory/asset_units.html', item=item, units=units,
                           warehouses=warehouses, sites=sites,
                           warehouse_list=warehouse_list, site_list=site_list,
                           counts=counts,
                           ASSET_STATUSES=ASSET_STATUSES, ASSET_CONDITIONS=ASSET_CONDITIONS)


@main.route('/items/<int:item_id>/units/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_inventory')
def asset_unit_create(item_id):
    item = Item.query.get_or_404(item_id)
    if not item.is_asset:
        abort(400)
    form = AssetUnitForm()
    if form.validate_on_submit():
        # Determine location type from chosen location id
        loc_type = form.location_type.data
        loc_id = form.location_id.data
        unit = AssetUnit(
            item_id=item_id,
            asset_tag=form.asset_tag.data.strip(),
            serial_number=form.serial_number.data or None,
            status=form.status.data,
            condition=form.condition.data,
            location_type=loc_type,
            location_id=loc_id,
            acquired_date=form.acquired_date.data,
            notes=form.notes.data or None,
        )
        db.session.add(unit)
        db.session.commit()
        flash(f'Unit {unit.asset_tag} registered.', 'success')
        if 'add_another' in request.form:
            return redirect(url_for('main.asset_unit_create', item_id=item_id))
        return redirect(url_for('main.asset_unit_list', item_id=item_id))
    # Auto-suggest tag
    last = AssetUnit.query.filter_by(item_id=item_id).order_by(desc(AssetUnit.id)).first()
    if last:
        parts = last.asset_tag.rsplit('-', 1)
        try:
            next_tag = f"{parts[0]}-{int(parts[1])+1:03d}"
        except (ValueError, IndexError):
            next_tag = f"{item.sku}-001"
    else:
        next_tag = f"{item.sku}-001"
    return render_template('inventory/asset_unit_form.html', form=form, item=item,
                           title='Register Asset Unit', suggested_tag=next_tag)


@main.route('/asset-units/<int:unit_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('manage_inventory')
def asset_unit_edit(unit_id):
    unit = AssetUnit.query.get_or_404(unit_id)
    item = unit.item
    form = AssetUnitForm(obj=unit)
    if form.validate_on_submit():
        unit.asset_tag = form.asset_tag.data.strip()
        unit.serial_number = form.serial_number.data or None
        unit.status = form.status.data
        unit.condition = form.condition.data
        unit.location_type = form.location_type.data
        unit.location_id = form.location_id.data
        unit.acquired_date = form.acquired_date.data
        unit.notes = form.notes.data or None
        db.session.commit()
        flash(f'Unit {unit.asset_tag} updated.', 'success')
        return redirect(url_for('main.asset_unit_list', item_id=item.id))
    return render_template('inventory/asset_unit_form.html', form=form, item=item,
                           unit=unit, title=f'Edit Unit — {unit.asset_tag}', suggested_tag=unit.asset_tag)


@main.route('/items/<int:item_id>/units/bulk-register', methods=['POST'])
@login_required
@permission_required('manage_inventory')
def asset_unit_bulk_register(item_id):
    """Quick-register multiple units with auto-incrementing tags."""
    item = Item.query.get_or_404(item_id)
    if not item.is_asset:
        abort(400)
    prefix = request.form.get('prefix', item.sku).strip()
    start_num = int(request.form.get('start_num', 1))
    count = min(int(request.form.get('count', 1)), 100)
    loc_type = request.form.get('location_type', 'warehouse')
    loc_id = int(request.form.get('location_id', 0))
    condition = request.form.get('condition', 'good')

    if not loc_id:
        flash('Please select a location.', 'danger')
        return redirect(url_for('main.asset_unit_list', item_id=item_id))

    created = 0
    for i in range(count):
        tag = f"{prefix}-{start_num + i:03d}"
        if not AssetUnit.query.filter_by(asset_tag=tag).first():
            db.session.add(AssetUnit(
                item_id=item_id, asset_tag=tag, status='available',
                condition=condition, location_type=loc_type, location_id=loc_id,
            ))
            created += 1
    db.session.commit()
    flash(f'{created} units registered successfully.', 'success')
    return redirect(url_for('main.asset_unit_list', item_id=item_id))


@main.route('/items/<int:item_id>/units/move', methods=['GET', 'POST'])
@login_required
@permission_required('manage_movements')
def asset_unit_move(item_id):
    """Move selected asset units to a new location."""
    item = Item.query.get_or_404(item_id)
    if not item.is_asset:
        abort(400)

    # Filter units by source location (from query params on GET)
    src_type = request.args.get('src_type', 'warehouse')
    src_id = int(request.args.get('src_id', 0))

    form = AssetMovementForm()

    if form.validate_on_submit():
        unit_ids = request.form.getlist('unit_ids')
        if not unit_ids:
            flash('Please select at least one unit to move.', 'danger')
            return redirect(url_for('main.asset_unit_move', item_id=item_id))

        mtype = form.movement_type.data
        to_loc_type = form.to_location_type.data
        to_loc_id = form.to_warehouse_id.data if to_loc_type == 'warehouse' else form.to_site_id.data

        # Determine new status
        status_map = {
            'delivery': 'available',      # Input: External → Warehouse
            'transfer': 'deployed',       # Delivery: Warehouse → Site
            'site_transfer': 'deployed',  # Transfer: Site → Site (still deployed)
            'pullout': 'available',       # Pullout: Site → Warehouse
            'maintenance': 'maintenance',
            'scrap': 'scrapped',
            'adjustment': None,           # keep current
        }
        new_status = status_map.get(mtype)

        # Create one Movement record for the batch
        from_wh_id = src_id if src_type == 'warehouse' else None
        from_site_id = src_id if src_type == 'site' else None
        to_wh_id = to_loc_id if to_loc_type == 'warehouse' else None
        to_site_id_val = to_loc_id if to_loc_type == 'site' else None

        movement = Movement(
            movement_type=mtype,
            item_id=item_id,
            quantity=len(unit_ids),
            unit_cost=item.unit_cost,
            from_location_type=src_type if src_id else 'external',
            from_warehouse_id=from_wh_id,
            from_site_id=from_site_id,
            to_location_type=to_loc_type if to_loc_type != 'none' else 'external',
            to_warehouse_id=to_wh_id,
            to_site_id=to_site_id_val,
            reference=form.reference.data or None,
            notes=form.notes.data or None,
            user_id=current_user.id,
        )
        db.session.add(movement)
        db.session.flush()

        moved = 0
        for uid in unit_ids:
            unit = AssetUnit.query.get(int(uid))
            if unit and unit.item_id == item_id:
                if new_status:
                    unit.status = new_status
                unit.condition = form.condition.data or unit.condition
                if to_loc_type != 'none' and to_loc_id:
                    unit.location_type = to_loc_type
                    unit.location_id = to_loc_id
                db.session.add(AssetUnitMovement(asset_unit_id=unit.id, movement_id=movement.id))
                moved += 1

        movement.quantity = moved
        db.session.commit()
        flash(f'{moved} unit(s) moved successfully.', 'success')
        return redirect(url_for('main.asset_unit_list', item_id=item_id))

    # Get available units at source for display
    if src_id:
        units = AssetUnit.query.filter_by(item_id=item_id, location_type=src_type,
                                          location_id=src_id).filter(
            AssetUnit.status != 'scrapped').order_by(AssetUnit.asset_tag).all()
    else:
        units = AssetUnit.query.filter_by(item_id=item_id).filter(
            AssetUnit.status != 'scrapped').order_by(AssetUnit.asset_tag).all()

    warehouses = Warehouse.query.filter_by(is_active=True).order_by('name').all()
    sites = ProjectSite.query.order_by('name').all()
    return render_template('inventory/asset_move.html', item=item, form=form, units=units,
                           warehouses=warehouses, sites=sites,
                           src_type=src_type, src_id=src_id,
                           ASSET_STATUSES=ASSET_STATUSES)


@main.route('/asset-units/api/by-item/<int:item_id>')
@login_required
def api_asset_units(item_id):
    """JSON API: return units for a given item, optionally filtered by location."""
    loc_type = request.args.get('loc_type')
    loc_id = request.args.get('loc_id', type=int)
    q = AssetUnit.query.filter_by(item_id=item_id).filter(AssetUnit.status != 'scrapped')
    if loc_type:
        q = q.filter_by(location_type=loc_type)
    if loc_id:
        q = q.filter_by(location_id=loc_id)
    return jsonify([{
        'id': u.id, 'tag': u.asset_tag, 'serial': u.serial_number or '',
        'status': u.status, 'status_label': u.status_label, 'status_color': u.status_color,
        'condition': u.condition, 'condition_label': u.condition_label,
        'location_name': u.location_name(),
    } for q_u in [q.order_by(AssetUnit.asset_tag).all()] for u in q_u])


# ─── ITEMS BULK IMPORT ────────────────────────────────────────────────────────

ALLOWED_IMPORT_EXTS = {'xlsx', 'xls', 'csv'}

def _allowed_import(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMPORT_EXTS


@main.route('/items/import/template')
@login_required
@permission_required('manage_inventory')
def item_import_template():
    """Download a blank Excel template for bulk item import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Items Import'

    headers = ['SKU*', 'Name*', 'Category', 'Unit*', 'Unit Cost', 'Reorder Level', 'Description']
    col_widths = [18, 40, 25, 12, 14, 16, 45]

    header_fill = PatternFill(start_color='1a6fd4', end_color='1a6fd4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Sample rows
    cats = [c.name for c in Category.query.order_by('name').limit(5).all()]
    samples = [
        ['TOOL-001', 'Safety Helmet (White)', cats[0] if cats else 'Safety Gear', 'pcs', 350.00, 20, 'Standard white hard hat'],
        ['TOOL-002', 'Work Gloves (Leather)', cats[0] if cats else 'Safety Gear', 'pairs', 120.00, 30, 'Heavy duty leather gloves'],
        ['MAT-001', 'Portland Cement', cats[1] if len(cats) > 1 else 'Materials', 'bags', 280.00, 50, '40kg bag Portland Type I'],
        ['MAT-002', 'Steel Rebar 10mm x 6m', cats[1] if len(cats) > 1 else 'Materials', 'pcs', 180.00, 100, 'Grade 60 deformed bar'],
        ['ELEC-001', 'THHN Wire 3.5mm', cats[2] if len(cats) > 2 else 'Electrical', 'm', 75.00, 200, 'THHN/THWN stranded wire'],
    ]
    alt_fill = PatternFill(start_color='EEF4FF', end_color='EEF4FF', fill_type='solid')

    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Instructions sheet
    ws_info = wb.create_sheet('Instructions')
    ws_info.column_dimensions['A'].width = 60
    notes = [
        ('PhilAsia Pro — Items Import Template', True),
        ('', False),
        ('COLUMN GUIDE:', True),
        ('SKU* — Unique item code (required). Will be skipped if already exists.', False),
        ('Name* — Item display name (required).', False),
        ('Category — Must match an existing category name exactly. Leave blank for none.', False),
        ('Unit* — Unit of measure (pcs, kg, bags, m, liters, etc.) — required.', False),
        ('Unit Cost — Cost per unit in Philippine Peso. Default 0.', False),
        ('Reorder Level — Low stock alert threshold. Default 0.', False),
        ('Description — Optional. Additional details about the item.', False),
        ('', False),
        ('TIPS:', True),
        ('• Rows with duplicate SKUs are skipped (existing items not overwritten).', False),
        ('• Rows missing SKU, Name, or Unit are skipped.', False),
        ('• Import supports .xlsx, .xls, and .csv formats.', False),
        ('• A summary of created / skipped rows is shown after import.', False),
    ]
    title_fill = PatternFill(start_color='1a6fd4', end_color='1a6fd4', fill_type='solid')
    for r, (text, bold) in enumerate(notes, 1):
        cell = ws_info.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, color='FFFFFF' if r == 1 else '111111', size=11 if bold else 10)
        if r == 1:
            cell.fill = title_fill
            ws_info.row_dimensions[r].height = 26

    # Freeze header row on import sheet
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='philasia_items_import_template.xlsx', as_attachment=True)


@main.route('/items/import', methods=['GET', 'POST'])
@login_required
@permission_required('manage_inventory')
def item_import():
    categories = Category.query.order_by('name').all()

    if request.method == 'POST':
        file = request.files.get('import_file')
        if not file or not file.filename:
            flash('Please select a file to import.', 'danger')
            return redirect(url_for('main.item_import'))
        if not _allowed_import(file.filename):
            flash('Unsupported file format. Please upload .xlsx, .xls, or .csv.', 'danger')
            return redirect(url_for('main.item_import'))

        ext = file.filename.rsplit('.', 1)[1].lower()
        created, skipped, errors = [], [], []

        try:
            if ext == 'csv':
                import csv, io
                stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
                reader = csv.DictReader(stream)
                rows = [r for r in reader]
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                raw_headers = [str(c.value).strip().rstrip('*') if c.value else '' for c in ws[1]]
                # Normalize headers
                header_map = {h.lower(): i for i, h in enumerate(raw_headers)}
                def get_col(row, name):
                    idx = header_map.get(name.lower())
                    return str(row[idx].value).strip() if idx is not None and row[idx].value is not None else ''
                rows = [{'sku': get_col(r, 'sku'), 'name': get_col(r, 'name'),
                         'category': get_col(r, 'category'), 'unit': get_col(r, 'unit'),
                         'unit cost': get_col(r, 'unit cost'), 'reorder level': get_col(r, 'reorder level'),
                         'description': get_col(r, 'description')}
                        for r in ws.iter_rows(min_row=2) if any(c.value for c in r)]

            cat_map = {c.name.lower(): c.id for c in categories}

            for i, row in enumerate(rows, 2):
                sku = (row.get('sku') or row.get('SKU') or row.get('SKU*') or '').strip()
                name = (row.get('name') or row.get('Name') or row.get('Name*') or '').strip()
                unit = (row.get('unit') or row.get('Unit') or row.get('Unit*') or '').strip()

                if not sku or not name or not unit:
                    skipped.append({'row': i, 'reason': 'Missing SKU, Name, or Unit', 'sku': sku or '(blank)'})
                    continue

                if Item.query.filter_by(sku=sku).first():
                    skipped.append({'row': i, 'reason': f'SKU already exists', 'sku': sku})
                    continue

                cat_name = (row.get('category') or row.get('Category') or '').strip().lower()
                cat_id = cat_map.get(cat_name)

                try:
                    unit_cost = float(row.get('unit cost') or row.get('Unit Cost') or row.get('unit_cost') or 0)
                except (ValueError, TypeError):
                    unit_cost = 0
                try:
                    reorder = float(row.get('reorder level') or row.get('Reorder Level') or row.get('reorder_level') or 0)
                except (ValueError, TypeError):
                    reorder = 0

                desc = (row.get('description') or row.get('Description') or '').strip()

                item = Item(sku=sku, name=name, unit=unit, unit_cost=unit_cost,
                            reorder_level=reorder, description=desc,
                            category_id=cat_id, is_active=True)
                db.session.add(item)
                created.append({'sku': sku, 'name': name})

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {str(e)}', 'danger')
            return redirect(url_for('main.item_import'))

        result = {'created': created, 'skipped': skipped}
        flash(f'{len(created)} items imported successfully. {len(skipped)} rows skipped.', 'success' if created else 'warning')
        return render_template('inventory/item_import_result.html', result=result)

    return render_template('inventory/item_import.html', categories=categories)


@main.route('/movements/vendor-input', methods=['GET', 'POST'])
@login_required
@permission_required('manage_movements')
def vendor_input():
    """Multi-item vendor input (purchase receipt).
    Allows recording a vendor delivery with multiple items in one go.
    New items not yet in the catalog can be created on the fly.
    """
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    categories = Category.query.order_by(Category.name).all()
    existing_items = Item.query.filter_by(is_active=True, item_type='consumable').order_by(Item.name).all()

    if request.method == 'POST':
        vendor_name = request.form.get('vendor_name', '').strip()
        reference = request.form.get('reference', '').strip()
        delivery_date = request.form.get('delivery_date', '').strip()
        dest_warehouse_id = int(request.form.get('dest_warehouse_id', 0) or 0)
        notes_global = request.form.get('notes', '').strip()

        if not dest_warehouse_id:
            flash('Please select a destination warehouse.', 'danger')
            return render_template('inventory/vendor_input.html',
                                   warehouses=warehouses, categories=categories,
                                   existing_items=existing_items, title='Vendor Input / Purchase Receipt')

        wh = Warehouse.query.get(dest_warehouse_id)
        if not wh:
            flash('Invalid warehouse selected.', 'danger')
            return render_template('inventory/vendor_input.html',
                                   warehouses=warehouses, categories=categories,
                                   existing_items=existing_items, title='Vendor Input / Purchase Receipt')

        # Parse the item rows (indexed by row number)
        row_indices = sorted(set(
            k.split('[')[1].split(']')[0]
            for k in request.form.keys()
            if k.startswith('rows[')
        ), key=lambda x: int(x))

        created_items = []
        movements_posted = []
        errors = []

        try:
            mov_date = datetime.strptime(delivery_date, '%Y-%m-%d') if delivery_date else datetime.utcnow()
        except ValueError:
            mov_date = datetime.utcnow()

        for idx in row_indices:
            row_type = request.form.get(f'rows[{idx}][row_type]', 'existing')
            qty_str = request.form.get(f'rows[{idx}][quantity]', '').strip()
            unit_cost_str = request.form.get(f'rows[{idx}][unit_cost]', '0').strip()

            try:
                qty = float(qty_str)
                if qty <= 0:
                    raise ValueError()
            except ValueError:
                errors.append(f'Row {int(idx)+1}: Invalid quantity "{qty_str}".')
                continue

            try:
                unit_cost = float(unit_cost_str or 0)
            except ValueError:
                unit_cost = 0

            item = None

            if row_type == 'existing':
                item_id = int(request.form.get(f'rows[{idx}][item_id]', 0) or 0)
                item = Item.query.get(item_id)
                if not item:
                    errors.append(f'Row {int(idx)+1}: No item selected.')
                    continue

            else:  # new item
                new_sku = request.form.get(f'rows[{idx}][new_sku]', '').strip().upper()
                new_name = request.form.get(f'rows[{idx}][new_name]', '').strip()
                new_unit = request.form.get(f'rows[{idx}][new_unit]', '').strip()
                new_item_type = request.form.get(f'rows[{idx}][new_item_type]', 'consumable')
                new_cat_id = int(request.form.get(f'rows[{idx}][new_category_id]', 0) or 0) or None

                if not new_name or not new_unit:
                    errors.append(f'Row {int(idx)+1}: New item must have a Name and Unit.')
                    continue

                # Auto-generate SKU if blank
                if not new_sku:
                    base = ''.join(c for c in new_name.upper()[:6] if c.isalnum())
                    candidate = base
                    counter = 1
                    while Item.query.filter_by(sku=candidate).first():
                        candidate = f"{base}-{counter:03d}"
                        counter += 1
                    new_sku = candidate

                # Check if SKU already exists — if so, use that item
                existing = Item.query.filter_by(sku=new_sku).first()
                if existing:
                    item = existing
                else:
                    item = Item(
                        sku=new_sku, name=new_name, unit=new_unit,
                        unit_cost=unit_cost, item_type=new_item_type,
                        category_id=new_cat_id, is_active=True
                    )
                    db.session.add(item)
                    db.session.flush()  # get item.id
                    created_items.append(f'{new_name} [{new_sku}]')

            # Build movement notes
            row_notes = []
            if vendor_name:
                row_notes.append(f'Vendor: {vendor_name}')
            if notes_global:
                row_notes.append(notes_global)

            movement = Movement(
                movement_type='delivery',
                item_id=item.id,
                quantity=qty,
                unit_cost=unit_cost or float(item.unit_cost or 0),
                from_location_type='external',
                to_location_type='warehouse',
                to_warehouse_id=dest_warehouse_id,
                reference=reference or None,
                notes=' | '.join(row_notes) if row_notes else None,
                user_id=current_user.id,
                date=mov_date,
            )
            db.session.add(movement)

            _update_stock(item.id, 'warehouse', dest_warehouse_id, None, qty)
            movements_posted.append({
                'name': item.name,
                'sku': item.sku,
                'qty': qty,
                'unit': item.unit,
                'unit_cost': unit_cost,
                'is_new': item.name in [n.split(' [')[0] for n in created_items],
            })

        if not movements_posted and not errors:
            flash('No valid items were submitted. Please add at least one row.', 'warning')
            return render_template('inventory/vendor_input.html',
                                   warehouses=warehouses, categories=categories,
                                   existing_items=existing_items, title='Vendor Input / Purchase Receipt')

        if movements_posted:
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                flash(f'Error saving: {str(e)}', 'danger')
                return render_template('inventory/vendor_input.html',
                                       warehouses=warehouses, categories=categories,
                                       existing_items=existing_items, title='Vendor Input / Purchase Receipt')

            for err in errors:
                flash(err, 'warning')
            return render_template('inventory/vendor_input_result.html',
                                   vendor_name=vendor_name,
                                   reference=reference,
                                   wh=wh,
                                   created_items=created_items,
                                   movements=movements_posted,
                                   title='Vendor Input — Receipt Confirmed')
        else:
            db.session.rollback()
            for err in errors:
                flash(err, 'danger')
            return render_template('inventory/vendor_input.html',
                                   warehouses=warehouses, categories=categories,
                                   existing_items=existing_items, title='Vendor Input / Purchase Receipt')

    import json as _json
    items_json = _json.dumps([
        {'id': i.id, 'name': i.name, 'sku': i.sku, 'unit': i.unit,
         'unit_cost': float(i.unit_cost or 0)}
        for i in existing_items
    ])
    cats_json = _json.dumps([{'id': c.id, 'name': c.name} for c in categories])
    return render_template('inventory/vendor_input.html',
                           warehouses=warehouses, categories=categories,
                           existing_items=existing_items,
                           existing_items_json=items_json,
                           categories_json=cats_json,
                           now=datetime.utcnow(),
                           title='Vendor Input / Purchase Receipt')


@main.route('/items/quick-add', methods=['POST'])
@login_required
@permission_required('manage_inventory')
def item_quick_add():
    """API endpoint for inline quick-add of multiple items."""
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({'success': False, 'error': 'Invalid data'}), 400

    cat_map = {c.name.lower(): c.id for c in Category.query.all()}
    created, errors = [], []

    for i, row in enumerate(data):
        sku = (row.get('sku') or '').strip()
        name = (row.get('name') or '').strip()
        unit = (row.get('unit') or '').strip()

        if not sku or not name or not unit:
            errors.append({'row': i + 1, 'error': 'SKU, Name, and Unit are required'})
            continue
        if Item.query.filter_by(sku=sku).first():
            errors.append({'row': i + 1, 'error': f'SKU "{sku}" already exists'})
            continue

        cat_name = (row.get('category') or '').strip().lower()
        try:
            item = Item(
                sku=sku, name=name, unit=unit,
                unit_cost=float(row.get('unit_cost') or 0),
                reorder_level=float(row.get('reorder_level') or 0),
                description=(row.get('description') or '').strip(),
                category_id=cat_map.get(cat_name),
                is_active=True
            )
            db.session.add(item)
            created.append({'sku': sku, 'name': name})
        except Exception as e:
            errors.append({'row': i + 1, 'error': str(e)})

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

    return jsonify({'success': True, 'created': created, 'errors': errors})


# ─── WAREHOUSES ───────────────────────────────────────────────────────────────

@main.route('/warehouses')
@login_required
def warehouse_list():
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    wh_data = []
    for wh in warehouses:
        items_count = db.session.query(func.count(Stock.id)).filter_by(
            warehouse_id=wh.id, location_type='warehouse').scalar() or 0
        total_value = db.session.query(func.sum(Stock.quantity * Item.unit_cost)).join(
            Item, Stock.item_id == Item.id).filter(
            Stock.warehouse_id == wh.id, Stock.location_type == 'warehouse').scalar() or 0
        wh_data.append({'wh': wh, 'items_count': items_count, 'total_value': float(total_value)})
    return render_template('inventory/warehouse_list.html', wh_data=wh_data)


@main.route('/warehouses/<int:wh_id>')
@login_required
def warehouse_detail(wh_id):
    wh = Warehouse.query.get_or_404(wh_id)
    stocks = Stock.query.filter_by(warehouse_id=wh_id, location_type='warehouse').all()
    return render_template('inventory/warehouse_detail.html', wh=wh, stocks=stocks)


@main.route('/warehouses/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_warehouses')
def warehouse_create():
    form = WarehouseForm()
    if form.validate_on_submit():
        wh = Warehouse(name=form.name.data, location=form.location.data,
                       contact_person=form.contact_person.data,
                       contact_info=form.contact_info.data,
                       is_active=form.is_active.data)
        db.session.add(wh)
        db.session.commit()
        flash('Warehouse registered.', 'success')
        return redirect(url_for('main.warehouse_list'))
    return render_template('inventory/warehouse_form.html', form=form, title='New Warehouse')


@main.route('/warehouses/<int:wh_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('manage_warehouses')
def warehouse_edit(wh_id):
    wh = Warehouse.query.get_or_404(wh_id)
    form = WarehouseForm(obj=wh)
    if form.validate_on_submit():
        wh.name = form.name.data
        wh.location = form.location.data
        wh.contact_person = form.contact_person.data
        wh.contact_info = form.contact_info.data
        wh.is_active = form.is_active.data
        db.session.commit()
        flash('Warehouse updated.', 'success')
        return redirect(url_for('main.warehouse_detail', wh_id=wh_id))
    return render_template('inventory/warehouse_form.html', form=form, title='Edit Warehouse', wh=wh)


# ─── PROJECT SITES ────────────────────────────────────────────────────────────

@main.route('/sites')
@login_required
def site_list():
    if current_user.role in ('admin', 'project_manager', 'accounting', 'finance_manager'):
        sites = ProjectSite.query.order_by(ProjectSite.name).all()
    else:
        sites = current_user.managed_sites
    site_data = []
    for site in sites:
        items_count = db.session.query(func.count(Stock.id)).filter_by(
            site_id=site.id, location_type='site').scalar() or 0
        pending_req = Request.query.filter_by(site_id=site.id, status='pending').count()
        site_data.append({'site': site, 'items_count': items_count, 'pending_req': pending_req})
    return render_template('inventory/site_list.html', site_data=site_data)


@main.route('/sites/<int:site_id>')
@login_required
def site_detail(site_id):
    site = ProjectSite.query.get_or_404(site_id)
    # Consumable stock at this site
    stocks = Stock.query.filter_by(site_id=site_id, location_type='site').filter(
        Stock.quantity > 0).all()
    # Filter to only consumable items
    consumable_stocks = [s for s in stocks if s.item.is_consumable]
    # Deployed asset units at this site
    deployed_assets = AssetUnit.query.filter_by(
        location_type='site', location_id=site_id
    ).filter(AssetUnit.status != 'scrapped').order_by(
        AssetUnit.item_id, AssetUnit.asset_tag).all()
    requests = Request.query.filter_by(site_id=site_id).order_by(desc(Request.created_at)).limit(10).all()
    movements = Movement.query.filter(
        (Movement.from_site_id == site_id) | (Movement.to_site_id == site_id)
    ).order_by(desc(Movement.date)).limit(15).all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    sites = ProjectSite.query.filter(ProjectSite.id != site_id).order_by(ProjectSite.name).all()
    return render_template('inventory/site_detail.html', site=site,
                           consumable_stocks=consumable_stocks,
                           deployed_assets=deployed_assets,
                           requests=requests, movements=movements,
                           warehouses=warehouses, other_sites=sites)


@main.route('/sites/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_sites')
def site_create():
    form = ProjectSiteForm()
    if form.validate_on_submit():
        site = ProjectSite(
            name=form.name.data, address=form.address.data, client=form.client.data,
            start_date=form.start_date.data, end_date=form.end_date.data,
            status=form.status.data, contact_person=form.contact_person.data,
            contact_phone=form.contact_phone.data,
            budget=form.budget.data or 0, notes=form.notes.data)
        db.session.add(site)
        db.session.commit()
        flash('Project site created.', 'success')
        return redirect(url_for('main.site_list'))
    return render_template('inventory/site_form.html', form=form, title='New Project Site')


@main.route('/sites/<int:site_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('manage_sites')
def site_edit(site_id):
    site = ProjectSite.query.get_or_404(site_id)
    form = ProjectSiteForm(obj=site)
    if form.validate_on_submit():
        site.name = form.name.data
        site.address = form.address.data
        site.client = form.client.data
        site.start_date = form.start_date.data
        site.end_date = form.end_date.data
        site.status = form.status.data
        site.contact_person = form.contact_person.data
        site.contact_phone = form.contact_phone.data
        site.budget = form.budget.data or 0
        site.notes = form.notes.data
        db.session.commit()
        flash('Project site updated.', 'success')
        return redirect(url_for('main.site_detail', site_id=site_id))
    return render_template('inventory/site_form.html', form=form, title='Edit Project Site', site=site)


# ─── MOVEMENTS ────────────────────────────────────────────────────────────────

@main.route('/movements')
@login_required
def movement_list():
    page = request.args.get('page', 1, type=int)
    mtype = request.args.get('type', '')
    q = Movement.query.order_by(desc(Movement.date))
    if mtype:
        q = q.filter_by(movement_type=mtype)
    movements = q.paginate(page=page, per_page=25)
    return render_template('inventory/movement_list.html', movements=movements,
                           mtype=mtype, movement_types=MOVEMENT_TYPES)


@main.route('/movements/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_movements')
def movement_create():
    form = MovementForm()
    if form.validate_on_submit():
        item = Item.query.get(form.item_id.data)
        if not item:
            flash('Item not found.', 'danger')
            return render_template('inventory/movement_form.html', form=form, title='New Material / Consumable Transaction')

        # Enforce: assets/tools/equipment must use the Asset Unit management page
        if item.is_asset:
            flash(
                f'"{item.name}" is a Tool / Equipment / Non-Consumable. '
                'Please use the Asset Unit Management page to transfer or pull out this item.',
                'warning'
            )
            return redirect(url_for('main.asset_unit_list', item_id=item.id))

        # Enforce: consumption is only valid for consumables (double guard)
        qty = float(form.quantity.data)
        mtype = form.movement_type.data
        from_type = form.from_location_type.data
        to_type = form.to_location_type.data
        from_wh = form.from_warehouse_id.data or None
        from_st = form.from_site_id.data or None
        to_wh = form.to_warehouse_id.data or None
        to_st = form.to_site_id.data or None

        # Validate source stock
        if from_type in ('warehouse', 'site'):
            q = Stock.query.filter_by(item_id=form.item_id.data, location_type=from_type)
            if from_type == 'warehouse':
                q = q.filter_by(warehouse_id=from_wh)
            else:
                q = q.filter_by(site_id=from_st)
            src_stock = q.first()
            if not src_stock or float(src_stock.quantity) < qty:
                flash('Insufficient stock at the source location.', 'danger')
                return render_template('inventory/movement_form.html', form=form, title='New Material / Consumable Transaction')

        movement = Movement(
            movement_type=mtype,
            item_id=form.item_id.data,
            quantity=qty,
            unit_cost=form.unit_cost.data or 0,
            from_location_type=from_type if from_type != 'none' else None,
            from_warehouse_id=from_wh if from_type == 'warehouse' else None,
            from_site_id=from_st if from_type == 'site' else None,
            to_location_type=to_type if to_type != 'none' else None,
            to_warehouse_id=to_wh if to_type == 'warehouse' else None,
            to_site_id=to_st if to_type == 'site' else None,
            reference=form.reference.data,
            notes=form.notes.data,
            user_id=current_user.id,
        )
        db.session.add(movement)

        # Update stock ledger
        _update_stock(form.item_id.data, from_type,
                      from_wh if from_type == 'warehouse' else None,
                      from_st if from_type == 'site' else None, -qty)
        _update_stock(form.item_id.data, to_type,
                      to_wh if to_type == 'warehouse' else None,
                      to_st if to_type == 'site' else None, qty)

        db.session.commit()
        flash('Transaction posted successfully.', 'success')
        return redirect(url_for('main.movement_list'))
    return render_template('inventory/movement_form.html', form=form, title='New Material / Consumable Transaction')


@main.route('/movements/<int:mov_id>')
@login_required
def movement_detail(mov_id):
    mov = Movement.query.get_or_404(mov_id)
    return render_template('inventory/movement_detail.html', mov=mov)


# ─── MATERIAL REQUESTS ────────────────────────────────────────────────────────

@main.route('/requests')
@login_required
def request_list():
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    q = Request.query.order_by(desc(Request.created_at))
    if status_filter:
        q = q.filter_by(status=status_filter)
    if current_user.role in ('project_manager',):
        # Show requests for sites they manage
        site_ids = [s.id for s in current_user.managed_sites]
        q = q.filter(Request.site_id.in_(site_ids))
    elif current_user.role == 'viewer':
        q = q.filter_by(user_id=current_user.id)
    reqs = q.paginate(page=page, per_page=20)
    return render_template('inventory/request_list.html', reqs=reqs, status_filter=status_filter)


@main.route('/requests/new', methods=['GET', 'POST'])
@login_required
def request_create():
    form = RequestForm()
    if form.validate_on_submit():
        req = Request(site_id=form.site_id.data, user_id=current_user.id,
                      date_needed=form.date_needed.data, priority=form.priority.data,
                      notes=form.notes.data)
        db.session.add(req)
        db.session.commit()
        flash('Requisition submitted. Add items below.', 'success')
        return redirect(url_for('main.request_detail', req_id=req.id))
    return render_template('inventory/request_form.html', form=form, title='New Material Requisition')


@main.route('/requests/<int:req_id>', methods=['GET', 'POST'])
@login_required
def request_detail(req_id):
    req = Request.query.get_or_404(req_id)
    item_form = RequestItemForm()
    approve_form = ApproveRequestForm()
    reject_form = RejectRequestForm()

    if item_form.validate_on_submit() and 'item_id' in request.form:
        if req.status == 'pending':
            ri = RequestItem(request_id=req_id, item_id=item_form.item_id.data,
                             quantity_requested=item_form.quantity_requested.data,
                             notes=item_form.notes.data)
            db.session.add(ri)
            db.session.commit()
            flash('Item added to requisition.', 'success')
        else:
            flash('Cannot modify an approved or fulfilled requisition.', 'warning')
        return redirect(url_for('main.request_detail', req_id=req_id))
    return render_template('inventory/request_detail.html', req=req, item_form=item_form,
                           approve_form=approve_form, reject_form=reject_form)


@main.route('/requests/<int:req_id>/approve', methods=['POST'])
@login_required
@permission_required('approve_requests')
def request_approve(req_id):
    req = Request.query.get_or_404(req_id)
    if req.status != 'pending':
        flash('This request has already been processed.', 'warning')
        return redirect(url_for('main.request_detail', req_id=req_id))
    req.status = 'approved'
    req.approved_by_id = current_user.id
    req.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Requisition approved.', 'success')
    return redirect(url_for('main.request_detail', req_id=req_id))


@main.route('/requests/<int:req_id>/reject', methods=['POST'])
@login_required
@permission_required('approve_requests')
def request_reject(req_id):
    req = Request.query.get_or_404(req_id)
    reason = request.form.get('rejection_reason', '')
    if req.status != 'pending':
        flash('This request has already been processed.', 'warning')
        return redirect(url_for('main.request_detail', req_id=req_id))
    req.status = 'rejected'
    req.rejection_reason = reason
    db.session.commit()
    flash('Requisition rejected.', 'warning')
    return redirect(url_for('main.request_detail', req_id=req_id))


@main.route('/requests/<int:req_id>/fulfill', methods=['POST'])
@login_required
@permission_required('manage_movements')
def request_fulfill(req_id):
    req = Request.query.get_or_404(req_id)
    if req.status not in ('approved', 'partial'):
        flash('Requisition must be approved before fulfillment.', 'warning')
        return redirect(url_for('main.request_detail', req_id=req_id))
    all_fulfilled = True
    for ri in req.items:
        remaining = float(ri.quantity_requested) - float(ri.quantity_delivered or 0)
        if remaining <= 0:
            continue
        # Check available stock across all warehouses
        avail = db.session.query(func.sum(Stock.quantity)).filter_by(
            item_id=ri.item_id, location_type='warehouse').scalar() or 0
        deliver_qty = min(float(avail), remaining)
        if deliver_qty <= 0:
            all_fulfilled = False
            continue
        # Find warehouse with stock
        src = Stock.query.filter_by(item_id=ri.item_id, location_type='warehouse').filter(
            Stock.quantity > 0).first()
        if not src:
            all_fulfilled = False
            continue
        mov = Movement(
            movement_type='transfer',
            item_id=ri.item_id, quantity=deliver_qty,
            from_location_type='warehouse', from_warehouse_id=src.warehouse_id,
            to_location_type='site', to_site_id=req.site_id,
            reference=f'REQ-{req.id}', notes=f'Fulfillment of Requisition #{req.id}',
            user_id=current_user.id, request_id=req_id)
        db.session.add(mov)
        _update_stock(ri.item_id, 'warehouse', src.warehouse_id, None, -deliver_qty)
        _update_stock(ri.item_id, 'site', None, req.site_id, deliver_qty)
        ri.quantity_delivered = float(ri.quantity_delivered or 0) + deliver_qty
        if float(ri.quantity_delivered) < float(ri.quantity_requested):
            all_fulfilled = False
    req.status = 'fulfilled' if all_fulfilled else 'partial'
    db.session.commit()
    flash('Requisition fulfilled.' if all_fulfilled else 'Partial fulfillment. Some items had insufficient stock.', 'success' if all_fulfilled else 'warning')
    return redirect(url_for('main.request_detail', req_id=req_id))


# ─── REPORTS ──────────────────────────────────────────────────────────────────

@main.route('/reports')
@login_required
def report_list():
    return render_template('reports/report_list.html')


@main.route('/reports/stock')
@login_required
def report_stock():
    form = ReportFilterForm(request.args, meta={'csrf': False})
    items = Item.query.filter_by(is_active=True).order_by(Item.name).all()
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    sites = ProjectSite.query.all()

    stock_data = []
    for item in items:
        row = {'item': item, 'warehouses': {}, 'sites': {}, 'total': 0}
        for wh in warehouses:
            s = Stock.query.filter_by(item_id=item.id, warehouse_id=wh.id, location_type='warehouse').first()
            qty = float(s.quantity) if s else 0
            row['warehouses'][wh.id] = qty
            row['total'] += qty
        for site in sites:
            s = Stock.query.filter_by(item_id=item.id, site_id=site.id, location_type='site').first()
            qty = float(s.quantity) if s else 0
            row['sites'][site.id] = qty
            row['total'] += qty
        stock_data.append(row)
    return render_template('reports/stock_report.html', stock_data=stock_data,
                           warehouses=warehouses, sites=sites, items=items)


@main.route('/reports/movements')
@login_required
def report_movements():
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    mtype = request.args.get('type', '')
    page = request.args.get('page', 1, type=int)

    q = Movement.query.order_by(desc(Movement.date))
    if date_from:
        try:
            q = q.filter(Movement.date >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(Movement.date <= datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59))
        except ValueError:
            pass
    if mtype:
        q = q.filter_by(movement_type=mtype)

    total_qty = q.with_entities(func.sum(Movement.quantity)).order_by(None).scalar() or 0
    movements = q.paginate(page=page, per_page=30)
    return render_template('reports/movements_report.html', movements=movements,
                           movement_types=MOVEMENT_TYPES, mtype=mtype,
                           date_from=date_from, date_to=date_to, total_qty=float(total_qty))


@main.route('/reports/low-stock')
@login_required
def report_low_stock():
    alerts = []
    for item in Item.query.filter_by(is_active=True).all():
        total = item.total_stock()
        if total < float(item.reorder_level or 0):
            alerts.append({'item': item, 'current': total, 'threshold': float(item.reorder_level or 0)})
    alerts.sort(key=lambda x: x['current'] / max(x['threshold'], 0.001))
    return render_template('reports/low_stock_report.html', alerts=alerts)


# ─── USER MANAGEMENT ─────────────────────────────────────────────────────────

@main.route('/users')
@login_required
@permission_required('manage_users')
def user_list():
    users = User.query.order_by(User.username).all()
    return render_template('users/user_list.html', users=users)


@main.route('/users/new', methods=['GET', 'POST'])
@login_required
@permission_required('manage_users')
def user_create():
    form = UserForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash('Username already taken.', 'danger')
            return render_template('users/user_form.html', form=form, title='New User')
        user = User(username=form.username.data, email=form.email.data,
                    full_name=form.full_name.data, role=form.role.data,
                    is_active=form.is_active.data)
        if form.password.data:
            user.set_password(form.password.data)
        else:
            user.set_password('password123')
        db.session.add(user)
        db.session.commit()
        flash(f'User {user.username} created.', 'success')
        return redirect(url_for('main.user_list'))
    return render_template('users/user_form.html', form=form, title='New User')


@main.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('manage_users')
def user_edit(user_id):
    user = User.query.get_or_404(user_id)
    form = UserForm(obj=user)
    if form.validate_on_submit():
        existing = User.query.filter_by(username=form.username.data).first()
        if existing and existing.id != user_id:
            flash('Username already taken.', 'danger')
            return render_template('users/user_form.html', form=form, title='Edit User', user=user)
        user.username = form.username.data
        user.email = form.email.data
        user.full_name = form.full_name.data
        user.role = form.role.data
        user.is_active = form.is_active.data
        if form.password.data:
            user.set_password(form.password.data)
        db.session.commit()
        flash('User updated.', 'success')
        return redirect(url_for('main.user_list'))
    return render_template('users/user_form.html', form=form, title='Edit User', user=user)


@main.route('/users/<int:user_id>/deactivate', methods=['POST'])
@login_required
@permission_required('manage_users')
def user_deactivate(user_id):
    if user_id == current_user.id:
        flash('You cannot deactivate your own account.', 'danger')
        return redirect(url_for('main.user_list'))
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    flash(f'User {"activated" if user.is_active else "deactivated"}.', 'success')
    return redirect(url_for('main.user_list'))


@main.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if not current_user.check_password(form.current_password.data):
            flash('Current password is incorrect.', 'danger')
            return render_template('users/profile.html', form=form)
        current_user.set_password(form.new_password.data)
        db.session.commit()
        flash('Password changed successfully.', 'success')
        return redirect(url_for('main.profile'))
    return render_template('users/profile.html', form=form)


# ─── INVENTORY OVERVIEW (Stocks) ──────────────────────────────────────────────

@main.route('/inventory')
@login_required
def inventory_overview():
    stocks = Stock.query.filter(Stock.quantity > 0).all()
    return render_template('inventory/inventory_overview.html', stocks=stocks)


# ─── API (JSON) ───────────────────────────────────────────────────────────────

@main.route('/api/stock/<int:item_id>')
@login_required
def api_item_stock(item_id):
    stocks = Stock.query.filter_by(item_id=item_id).all()
    return jsonify([{
        'location_type': s.location_type,
        'location': s.location_name(),
        'quantity': float(s.quantity),
    } for s in stocks])


# ─── GLOBAL SEARCH ────────────────────────────────────────────────────────────

@main.route('/api/search')
@login_required
def api_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    pat = f'%{q}%'
    results = []

    items = Item.query.filter(
        Item.is_active == True,
        or_(Item.name.ilike(pat), Item.sku.ilike(pat), Item.description.ilike(pat))
    ).limit(6).all()
    for i in items:
        results.append({'type': 'item', 'label': i.name, 'sub': i.sku,
                        'url': url_for('main.item_detail', item_id=i.id),
                        'icon': 'box-seam'})

    assets = AssetUnit.query.filter(
        or_(AssetUnit.asset_tag.ilike(pat), AssetUnit.serial_number.ilike(pat))
    ).limit(5).all()
    for a in assets:
        results.append({'type': 'asset', 'label': a.asset_tag,
                        'sub': a.item.name,
                        'url': url_for('main.asset_unit_list', item_id=a.item_id),
                        'icon': 'tag'})

    sites = ProjectSite.query.filter(
        or_(ProjectSite.name.ilike(pat), ProjectSite.client.ilike(pat))
    ).limit(4).all()
    for s in sites:
        results.append({'type': 'site', 'label': s.name, 'sub': s.client,
                        'url': url_for('main.site_detail', site_id=s.id),
                        'icon': 'geo-alt'})

    warehouses = Warehouse.query.filter(
        Warehouse.name.ilike(pat)
    ).limit(3).all()
    for w in warehouses:
        results.append({'type': 'warehouse', 'label': w.name, 'sub': w.location or '',
                        'url': url_for('main.warehouse_detail', wh_id=w.id),
                        'icon': 'building'})

    return jsonify(results[:15])


# ─── QUICK CONSUMPTION ────────────────────────────────────────────────────────

@main.route('/quick/consume', methods=['GET', 'POST'])
@login_required
@permission_required('manage_movements')
def quick_consume():
    sites = ProjectSite.query.filter(ProjectSite.status.in_(['active', 'planned'])).order_by('name').all()
    items_consumable = Item.query.filter_by(is_active=True, item_type='consumable').order_by('name').all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by('name').all()

    if request.method == 'POST':
        action = request.form.get('action', 'consumption')
        item_id = int(request.form.get('item_id', 0))
        qty = float(request.form.get('quantity', 0))
        site_id = int(request.form.get('site_id', 0)) or None
        warehouse_id = int(request.form.get('warehouse_id', 0)) or None
        notes = request.form.get('notes', '')
        reference = request.form.get('reference', '')

        item = Item.query.get(item_id)
        if not item or qty <= 0:
            flash('Please select a valid item and quantity.', 'danger')
            return render_template('inventory/quick_consume.html', sites=sites,
                                   items=items_consumable, warehouses=warehouses)

        if action == 'consumption':
            if not site_id:
                flash('Please select a site for consumption.', 'danger')
                return render_template('inventory/quick_consume.html', sites=sites,
                                       items=items_consumable, warehouses=warehouses)
            src = Stock.query.filter_by(item_id=item_id, location_type='site', site_id=site_id).first()
            avail = float(src.quantity) if src else 0
            if avail < qty:
                flash(f'Insufficient stock at site. Available: {avail} {item.unit}', 'danger')
                return render_template('inventory/quick_consume.html', sites=sites,
                                       items=items_consumable, warehouses=warehouses)
            mov = Movement(movement_type='consumption', item_id=item_id, quantity=qty,
                           unit_cost=item.unit_cost,
                           from_location_type='site', from_site_id=site_id,
                           to_location_type=None, reference=reference, notes=notes,
                           user_id=current_user.id)
            db.session.add(mov)
            _update_stock(item_id, 'site', None, site_id, -qty)

        elif action == 'pullout':
            if not site_id:
                flash('Please select the source site.', 'danger')
                return render_template('inventory/quick_consume.html', sites=sites,
                                       items=items_consumable, warehouses=warehouses)
            if not warehouse_id:
                flash('Please select the destination warehouse.', 'danger')
                return render_template('inventory/quick_consume.html', sites=sites,
                                       items=items_consumable, warehouses=warehouses)
            src = Stock.query.filter_by(item_id=item_id, location_type='site', site_id=site_id).first()
            avail = float(src.quantity) if src else 0
            if avail < qty:
                flash(f'Insufficient stock at site. Available: {avail} {item.unit}', 'danger')
                return render_template('inventory/quick_consume.html', sites=sites,
                                       items=items_consumable, warehouses=warehouses)
            mov = Movement(movement_type='pullout', item_id=item_id, quantity=qty,
                           unit_cost=item.unit_cost,
                           from_location_type='site', from_site_id=site_id,
                           to_location_type='warehouse', to_warehouse_id=warehouse_id,
                           reference=reference, notes=notes, user_id=current_user.id)
            db.session.add(mov)
            _update_stock(item_id, 'site', None, site_id, -qty)
            _update_stock(item_id, 'warehouse', warehouse_id, None, qty)

        elif action == 'adjustment':
            loc_type = request.form.get('adj_loc_type', 'warehouse')
            loc_id_adj = int(request.form.get('adj_loc_id', 0)) or None
            if not loc_id_adj:
                flash('Please select a location for adjustment.', 'danger')
                return render_template('inventory/quick_consume.html', sites=sites,
                                       items=items_consumable, warehouses=warehouses)
            new_qty = float(request.form.get('new_quantity', 0))
            # Find current stock
            q_filter = Stock.query.filter_by(item_id=item_id, location_type=loc_type)
            if loc_type == 'warehouse':
                q_filter = q_filter.filter_by(warehouse_id=loc_id_adj)
            else:
                q_filter = q_filter.filter_by(site_id=loc_id_adj)
            cur_stock = q_filter.first()
            cur_qty = float(cur_stock.quantity) if cur_stock else 0
            delta = new_qty - cur_qty
            if delta == 0:
                flash('No change — stock quantity is already at that level.', 'info')
                return redirect(url_for('main.quick_consume'))
            mov = Movement(movement_type='adjustment', item_id=item_id, quantity=abs(delta),
                           unit_cost=item.unit_cost,
                           from_location_type=loc_type if delta < 0 else None,
                           from_warehouse_id=loc_id_adj if loc_type == 'warehouse' and delta < 0 else None,
                           from_site_id=loc_id_adj if loc_type == 'site' and delta < 0 else None,
                           to_location_type=loc_type if delta > 0 else None,
                           to_warehouse_id=loc_id_adj if loc_type == 'warehouse' and delta > 0 else None,
                           to_site_id=loc_id_adj if loc_type == 'site' and delta > 0 else None,
                           reference=reference, notes=notes or f'Stock adjustment: {cur_qty} → {new_qty}',
                           user_id=current_user.id)
            db.session.add(mov)
            _update_stock(item_id, loc_type,
                          loc_id_adj if loc_type == 'warehouse' else None,
                          loc_id_adj if loc_type == 'site' else None,
                          delta)

        db.session.commit()
        flash(f'Transaction recorded: {qty} {item.unit} of {item.name}.', 'success')
        return redirect(url_for('main.quick_consume'))

    prefill_item = request.args.get('item_id', type=int)
    prefill_site = request.args.get('site_id', type=int)
    recent_movs = Movement.query.filter(
        Movement.movement_type.in_(['consumption', 'pullout', 'adjustment'])
    ).order_by(desc(Movement.date)).limit(10).all()
    return render_template('inventory/quick_consume.html', sites=sites,
                           items=items_consumable, warehouses=warehouses,
                           prefill_item=prefill_item, prefill_site=prefill_site,
                           recent_movs=recent_movs)


# ─── QR CODE GENERATION ───────────────────────────────────────────────────────

@main.route('/qr/item/<int:item_id>')
@login_required
def qr_item(item_id):
    item = Item.query.get_or_404(item_id)
    import qrcode
    url = url_for('main.item_detail', item_id=item_id, _external=True)
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png',
                     download_name=f'qr_{item.sku}.png')


@main.route('/qr/asset/<int:unit_id>')
@login_required
def qr_asset_unit(unit_id):
    unit = AssetUnit.query.get_or_404(unit_id)
    import qrcode
    url = url_for('main.asset_unit_list', item_id=unit.item_id, _external=True)
    data = f"TAG:{unit.asset_tag}|ITEM:{unit.item.name}|SN:{unit.serial_number or '-'}|URL:{url}"
    img = qrcode.make(data)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png',
                     download_name=f'qr_{unit.asset_tag}.png')


@main.route('/items/<int:item_id>/labels')
@login_required
def item_labels(item_id):
    item = Item.query.get_or_404(item_id)
    units = []
    if item.is_asset:
        units = AssetUnit.query.filter_by(item_id=item_id).order_by(AssetUnit.asset_tag).all()
    warehouses = {w.id: w.name for w in Warehouse.query.all()}
    sites = {s.id: s.name for s in ProjectSite.query.all()}
    return render_template('inventory/item_labels.html', item=item, units=units,
                           warehouses=warehouses, sites=sites)


# ─── EXCEL EXPORTS ────────────────────────────────────────────────────────────

def _make_workbook():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='1a6fd4', end_color='1a6fd4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color='EEF4FF', end_color='EEF4FF', fill_type='solid')
    return wb, header_fill, header_font, border, alt_fill


@main.route('/reports/stock/export')
@login_required
def export_stock():
    wb, hfill, hfont, border, altfill = _make_workbook()
    from openpyxl.styles import Alignment
    ws = wb.active
    ws.title = 'Stock Report'

    items = Item.query.filter_by(is_active=True).order_by(Item.name).all()
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    sites = ProjectSite.query.all()

    headers = ['SKU', 'Item Name', 'Type', 'Unit', 'Reorder Level', 'Total Stock', 'Status']
    for w in warehouses:
        headers.append(f'WH: {w.name}')
    for s in sites:
        headers.append(f'Site: {s.name}')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    for row_i, item in enumerate(items, 2):
        total = item.total_stock()
        status = 'LOW STOCK' if item.is_low_stock() else 'OK'
        row_data = [item.sku, item.name, item.type_label, item.unit,
                    float(item.reorder_level or 0), total, status]
        for w in warehouses:
            s = Stock.query.filter_by(item_id=item.id, warehouse_id=w.id, location_type='warehouse').first()
            row_data.append(float(s.quantity) if s else 0)
        for site in sites:
            s = Stock.query.filter_by(item_id=item.id, site_id=site.id, location_type='site').first()
            row_data.append(float(s.quantity) if s else 0)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border = border
            if row_i % 2 == 0:
                cell.fill = altfill

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'stock_report_{date.today()}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)


@main.route('/reports/movements/export')
@login_required
def export_movements():
    wb, hfill, hfont, border, altfill = _make_workbook()
    from openpyxl.styles import Alignment
    ws = wb.active
    ws.title = 'Movement Log'

    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    mtype = request.args.get('type', '')

    q = Movement.query.order_by(desc(Movement.date))
    if date_from:
        try:
            q = q.filter(Movement.date >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(Movement.date <= datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59))
        except ValueError:
            pass
    if mtype:
        q = q.filter_by(movement_type=mtype)

    movements = q.all()

    headers = ['#', 'Date', 'Type', 'Item', 'SKU', 'Quantity', 'Unit', 'Unit Cost',
               'Total Value', 'From', 'To', 'Reference', 'Notes', 'Posted By']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    for row_i, m in enumerate(movements, 2):
        total_val = float(m.quantity or 0) * float(m.unit_cost or 0)
        row_data = [m.id, m.date.strftime('%Y-%m-%d %H:%M'),
                    MOVEMENT_TYPES.get(m.movement_type, m.movement_type),
                    m.item.name, m.item.sku,
                    float(m.quantity), m.item.unit, float(m.unit_cost or 0), total_val,
                    m.from_location_name(), m.to_location_name(),
                    m.reference or '', m.notes or '', m.created_by.username]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border = border
            if row_i % 2 == 0:
                cell.fill = altfill

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'movements_{date.today()}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)


@main.route('/reports/low-stock/export')
@login_required
def export_low_stock():
    wb, hfill, hfont, border, altfill = _make_workbook()
    from openpyxl.styles import Alignment, Font, PatternFill
    ws = wb.active
    ws.title = 'Low Stock Alerts'

    headers = ['SKU', 'Item Name', 'Type', 'Unit', 'Current Stock', 'Reorder Level', 'Shortage']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    row_i = 2
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    for item in Item.query.filter_by(is_active=True).all():
        total = item.total_stock()
        if total < float(item.reorder_level or 0):
            shortage = float(item.reorder_level or 0) - total
            row_data = [item.sku, item.name, item.type_label, item.unit,
                        total, float(item.reorder_level or 0), shortage]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.border = border
                cell.fill = red_fill
            row_i += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'low_stock_{date.today()}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)


# ─── CATEGORY OVERVIEW ────────────────────────────────────────────────────────

@main.route('/categories/overview')
@login_required
def category_overview():
    categories = Category.query.order_by(Category.name).all()
    cat_data = []
    for cat in categories:
        items = Item.query.filter_by(category_id=cat.id, is_active=True).all()
        total_items = len(items)
        low_count = sum(1 for i in items if i.is_low_stock())
        total_value = sum(float(i.unit_cost or 0) * i.total_stock() for i in items)
        cat_data.append({
            'cat': cat, 'total_items': total_items,
            'low_count': low_count, 'total_value': total_value,
            'items': items
        })
    return render_template('inventory/category_overview.html', cat_data=cat_data)


# ─── SITE CONSUMPTION QUICK ENTRY ─────────────────────────────────────────────

@main.route('/sites/<int:site_id>/consume', methods=['POST'])
@login_required
@permission_required('manage_movements')
def site_quick_consume(site_id):
    site = ProjectSite.query.get_or_404(site_id)
    item_id = int(request.form.get('item_id', 0))
    qty = float(request.form.get('quantity', 0))
    action = request.form.get('action', 'consumption')

    item = Item.query.get(item_id)
    if not item or qty <= 0:
        flash('Invalid item or quantity.', 'danger')
        return redirect(url_for('main.site_detail', site_id=site_id))

    src = Stock.query.filter_by(item_id=item_id, location_type='site', site_id=site_id).first()
    avail = float(src.quantity) if src else 0
    if avail < qty:
        flash(f'Not enough stock. Available: {avail} {item.unit}', 'danger')
        return redirect(url_for('main.site_detail', site_id=site_id))

    if action == 'consumption':
        mov = Movement(movement_type='consumption', item_id=item_id, quantity=qty,
                       unit_cost=item.unit_cost,
                       from_location_type='site', from_site_id=site_id,
                       to_location_type=None,
                       notes=f'Quick consume from site page', user_id=current_user.id)
        db.session.add(mov)
        _update_stock(item_id, 'site', None, site_id, -qty)
        db.session.commit()
        flash(f'Recorded: {qty} {item.unit} of {item.name} consumed.', 'success')
    elif action == 'pullout':
        wh_id = int(request.form.get('warehouse_id', 0))
        wh = Warehouse.query.get(wh_id)
        if not wh:
            flash('Please select a warehouse for pullout.', 'danger')
            return redirect(url_for('main.site_detail', site_id=site_id))
        mov = Movement(movement_type='pullout', item_id=item_id, quantity=qty,
                       unit_cost=item.unit_cost,
                       from_location_type='site', from_site_id=site_id,
                       to_location_type='warehouse', to_warehouse_id=wh_id,
                       notes=f'Quick pullout from site page', user_id=current_user.id)
        db.session.add(mov)
        _update_stock(item_id, 'site', None, site_id, -qty)
        _update_stock(item_id, 'warehouse', wh_id, None, qty)
        db.session.commit()
        flash(f'Pulled out: {qty} {item.unit} of {item.name} → {wh.name}.', 'success')

    return redirect(url_for('main.site_detail', site_id=site_id))
